#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 转 Markdown 表格工具

将 Excel 格式的 Mapping 文档转换为 LLM 可读的 Markdown 表格格式
"""

import sys
import os

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("错误：需要安装 openpyxl 库")
    print("安装命令：pip install openpyxl")
    sys.exit(1)


def excel_to_markdown_table(worksheet, max_rows=None):
    """
    将 Excel 工作表转换为 Markdown 表格

    Args:
        worksheet: openpyxl 工作表对象
        max_rows: 最大行数（None 表示不限制）

    Returns:
        Markdown 表格字符串
    """
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return ""

    # 过滤空行
    rows = [row for row in rows if any(cell is not None for cell in row)]

    if max_rows:
        rows = rows[:max_rows]

    if not rows:
        return ""

    # 检测列数
    num_cols = max(len(row) for row in rows)

    # 构建表格
    lines = []
    header = rows[0]

    # 表头行
    header_cells = [str(cell) if cell is not None else '' for cell in header]
    # 补齐列数
    while len(header_cells) < num_cols:
        header_cells.append('')
    lines.append('| ' + ' | '.join(header_cells) + ' |')

    # 分隔线
    lines.append('| ' + ' | '.join(['---'] * num_cols) + ' |')

    # 数据行
    for row in rows[1:]:
        cells = [str(cell) if cell is not None else '' for cell in row]
        while len(cells) < num_cols:
            cells.append('')
        lines.append('| ' + ' | '.join(cells) + ' |')

    return '\n'.join(lines)


def excel_file_to_markdown(filepath, sheet_name=None, max_rows=None):
    """
    将 Excel 文件转换为 Markdown 表格

    Args:
        filepath: Excel 文件路径
        sheet_name: 工作表名称（None 表示第一个工作表）
        max_rows: 最大行数

    Returns:
        Markdown 表格字符串，如果有多个工作表则返回多个表格
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"文件不存在：{filepath}")

    wb = load_workbook(filename=filepath, read_only=True, data_only=True)

    results = []

    if sheet_name:
        # 指定工作表
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            md_table = excel_to_markdown_table(ws, max_rows)
            results.append(f"## {sheet_name}\n\n{md_table}")
    else:
        # 所有工作表
        for name in wb.sheetnames:
            ws = wb[name]
            md_table = excel_to_markdown_table(ws, max_rows)
            if md_table:
                results.append(f"## {name}\n\n{md_table}")

    wb.close()
    return '\n\n'.join(results)


def detect_table_type(content):
    """
    检测表格类型（表级 mapping 或 字段级 mapping）

    Args:
        content: Markdown 表格内容

    Returns:
        表格类型字符串
    """
    # 检查关键字
    if '来源表' in content or '目标表' in content or 'JOIN' in content:
        return "表级 Mapping"
    elif '目标字段' in content or '来源字段' in content or '转换规则' in content:
        return "字段级 Mapping"
    else:
        return "未知类型"


def main():
    """命令行入口"""
    if len(sys.argv) < 2:
        print("用法：python3 excel_to_markdown.py <excel_file> [sheet_name] [max_rows]")
        print("")
        print("参数说明:")
        print("  excel_file  - Excel 文件路径")
        print("  sheet_name  - 工作表名称（可选，默认为第一个工作表）")
        print("  max_rows    - 最大行数（可选，不限制则省略）")
        print("")
        print("示例:")
        print("  python3 excel_to_markdown.py mapping.xlsx")
        print("  python3 excel_to_markdown.py mapping.xlsx 表级 mapping")
        print("  python3 excel_to_markdown.py mapping.xlsx 字段级 mapping 100")
        sys.exit(1)

    filepath = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    max_rows = int(sys.argv[3]) if len(sys.argv) > 3 else None

    try:
        result = excel_file_to_markdown(filepath, sheet_name, max_rows)
        print(result)
        print("")
        print("# 表格类型检测：", detect_table_type(result))
    except Exception as e:
        print(f"错误：{e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
